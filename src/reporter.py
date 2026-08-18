import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_FAMILY = "Segoe UI"


def _base_styles():
    return {
        "header_font": Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF"),
        "title_font": Font(name=FONT_FAMILY, size=14, bold=True, color="1F2937"),
        "regular_font": Font(name=FONT_FAMILY, size=10),
        "bold_font": Font(name=FONT_FAMILY, size=10, bold=True),
        "header_fill": PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid"),  # Indigo
        "match_fill": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),    # Verde
        "diff_fill": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),     # Rojo
        "warning_fill": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),  # Amarillo
        "info_fill": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),     # Azul
        "thin_border": Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        ),
    }


def _autosize_columns(ws, header_row=1):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row >= header_row and cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min(max_len + 4, 60), 12)


def _write_category_sheet(wb, sheet_name, rows, columns, row_fill):
    """
    columns: lista de tuplas (clave_dict, encabezado) en el orden en que deben aparecer.
    row_fill: color aplicado a toda fila de datos -- una sola hoja = una sola categoría,
    así que (a diferencia del reporte de antes) no hace falta colorear fila por fila
    según el estado individual.
    """
    styles = _base_styles()
    ws = wb.create_sheet()
    ws.title = sheet_name

    headers = [label for _, label in columns]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows:
        ws.append(["Sin registros en esta categoría."])
        ws.cell(row=2, column=1).font = styles["regular_font"]
    else:
        for row_data in rows:
            values = [row_data.get(key, "") for key, _ in columns]
            ws.append(values)
            r_idx = ws.max_row
            for c_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = styles["regular_font"]
                cell.border = styles["thin_border"]
                cell.fill = row_fill

    _autosize_columns(ws)
    return ws


def _write_resumen_sheet(wb, key_col, compare_cols, metrics, run_meta):
    styles = _base_styles()
    ws = wb.active
    ws.title = "Resumen"

    ws.append(["Reporte de Auditoría OCR vs Excel"])
    ws.cell(row=1, column=1).font = styles["title_font"]
    ws.append([])

    run_meta = run_meta or {}

    elapsed = run_meta.get("elapsed_seconds")
    if elapsed is not None:
        minutos, segundos = divmod(round(elapsed), 60)
        tiempo_str = f"{minutos}m {segundos}s" if minutos else f"{segundos}s"
    else:
        tiempo_str = ""

    meta_rows = [
        ("Generado:", run_meta.get("generated_at", "")),
        ("Archivo Excel:", run_meta.get("excel_filename") or ""),
        ("Archivo PDF:", run_meta.get("pdf_filename") or ""),
        ("Páginas procesadas:", f"{run_meta.get('start_page', '')} - {run_meta.get('end_page', '')}"),
        ("Columna Llave (Documento):", str(key_col)),
        ("Columnas Comparadas:", ", ".join(compare_cols) if compare_cols else ""),
        ("Umbral de Similitud:", f"{run_meta.get('similarity_threshold', '')}%"),
        ("Tiempo de Procesamiento:", tiempo_str),
    ]
    for label, value in meta_rows:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(row=r, column=1).font = styles["regular_font"]
        ws.cell(row=r, column=2).font = styles["bold_font"]

    ws.append([])
    ws.append(["Resultados"])
    ws.cell(row=ws.max_row, column=1).font = styles["title_font"]
    ws.append([])

    tiles = [
        ("Verificados Perfectos", metrics.get("correct", 0), styles["match_fill"]),
        ("Alertas y Anomalías", metrics.get("alerts", 0), styles["diff_fill"]),
        ("Faltantes en PDF", metrics.get("missing", 0), styles["warning_fill"]),
        ("Solo en PDF (Huérfanos)", metrics.get("huerfanos", 0), styles["info_fill"]),
    ]
    header_row = ws.max_row + 1
    ws.append([t[0] for t in tiles])
    ws.append([t[1] for t in tiles])
    for c_idx, (_, _, fill) in enumerate(tiles, 1):
        header_cell = ws.cell(row=header_row, column=c_idx)
        header_cell.font = styles["bold_font"]
        header_cell.fill = fill
        header_cell.alignment = Alignment(horizontal="center")
        value_cell = ws.cell(row=header_row + 1, column=c_idx)
        value_cell.font = Font(name=FONT_FAMILY, size=18, bold=True, color="1F2937")
        value_cell.alignment = Alignment(horizontal="center")

    _autosize_columns(ws, header_row=1)
    return ws


def generate_excel_report(key_col, compare_cols, lista_coinciden, lista_anomalias,
                           lista_solo_pdf, lista_solo_excel, metrics, run_meta=None):
    """
    Genera el reporte de auditoría en 5 hojas: Resumen + una por categoría
    (Verificados Perfectos, Alertas y Anomalías, Solo en PDF (Huérfanos),
    Solo en Excel (Faltantes)). Reemplaza el reporte de una sola hoja de antes,
    que además nunca llegaba a mostrar huérfanos ni faltantes.
    """
    output = io.BytesIO()
    wb = Workbook()
    styles = _base_styles()

    _write_resumen_sheet(wb, key_col, compare_cols, metrics, run_meta)

    columnas_base_pdf = [
        ("Tipo_Documento_PDF", "Tipo Documento"),
        ("Fecha_Nacimiento_PDF", "Fecha Nacimiento"),
        ("Edad_PDF", "Edad"),
        ("Lugar_Nacimiento_PDF", "Lugar Nacimiento"),
        ("Sexo_PDF", "Sexo"),
        ("Estatura_PDF", "Estatura"),
        ("Grupo_Sanguineo_PDF", "RH"),
        ("Fecha_Lugar_Expedicion_PDF", "Fecha y Lugar Expedición"),
    ]

    columnas_comparacion = []
    for col in compare_cols:
        columnas_comparacion.append((f"{col}_excel", f"{col} (Excel)"))
        columnas_comparacion.append((f"{col}_score", f"{col} (Similitud %)"))

    columnas_coinciden_anomalias = (
        [("Página_PDF", "Página PDF"), ("Identificación_Excel", "Documento (Excel)"),
         ("Nombre_Excel", "Nombre (Excel)"), ("Identificación_PDF", "Documento (PDF)"),
         ("Nombre_PDF", "Nombre (PDF)")]
        + columnas_base_pdf
        + [("Similitud_Nombre_%", "Similitud Nombre %")]
        + columnas_comparacion
    )

    _write_category_sheet(
        wb, "Verificados Perfectos", lista_coinciden,
        columnas_coinciden_anomalias, styles["match_fill"]
    )
    _write_category_sheet(
        wb, "Alertas y Anomalías", lista_anomalias,
        columnas_coinciden_anomalias
        + [("Alerta_Detalle", "Detalle de Alerta"), ("Texto_Completo_PDF", "Texto Completo OCR")],
        styles["diff_fill"]
    )
    _write_category_sheet(
        wb, "Solo en PDF (Huérfanos)", lista_solo_pdf,
        [("Página_PDF", "Página PDF"), ("Identificación_PDF", "Documento (PDF)"), ("Nombre_PDF", "Nombre (PDF)")]
        + columnas_base_pdf + [("Texto_Completo_PDF", "Texto Completo OCR")],
        styles["warning_fill"]
    )
    _write_category_sheet(
        wb, "Solo en Excel (Faltantes)", lista_solo_excel,
        [("Identificación_Excel", "Documento (Excel)"), ("Nombre_Excel", "Nombre (Excel)")] + columnas_comparacion,
        styles["warning_fill"]
    )

    wb.save(output)
    return output.getvalue()
